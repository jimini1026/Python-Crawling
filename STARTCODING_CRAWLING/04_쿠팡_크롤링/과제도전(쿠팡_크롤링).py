import requests
from bs4 import BeautifulSoup
import time
import pyautogui
import openpyxl

wb = openpyxl.Workbook()

keyword = pyautogui.prompt("검색어를 입력하세요.")
page = 1
count = 0

ws = wb.create_sheet(f"{keyword} 100순위")
ws["A1"] = "순위"
ws["B1"] = "브랜드명"
ws["C1"] = "제품명"
ws["D1"] = "상세페이지"

while True:
    main_url = f"https://www.coupang.com/np/search?q={keyword}&page={page}&channel=recent"

    # 쿠팡에선 헤더에 User-Agent 추가하지 않으면 오류 발생(멈춰버림)
    # 헤더에 User-Agent, Accept-Language 를 추가하지 않으면 멈춥니다
    header = {
        'Host': 'www.coupang.com',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:76.0) Gecko/20100101 Firefox/76.0',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
        'Accept-Language': 'ko-KR,ko;q=0.8,en-US;q=0.5,en;q=0.3',
    }

    response = requests.get(main_url, headers = header)
    html = response.text
    soup = BeautifulSoup(html, "html.parser")

    links = soup.select("a.search-product-link") # select의 결과는 리스트

    for link in links:
        ad = link.select_one("span.ad-badge-text")
        if ad == None:
            if count != 100:
                sub_url = "https://www.coupang.com" + link.attrs["href"]

                sub_response = requests.get(sub_url, headers = header)
                sub_html = sub_response.text
                sub_soup = BeautifulSoup(sub_html, "html.parser")
                name = sub_soup.select_one("h2.prod-buy-header__title").text

                brand = sub_soup.select_one(".prod-brand-name").text.strip()
                count += 1
                ws[f"A{count + 1}"] = count
                ws[f"B{count + 1}"] = brand
                ws[f"C{count + 1}"] = name
                ws[f"D{count + 1}"] = sub_url
            else:
                break

        time.sleep(0.3)

    if count == 100:
        break

wb.save(f"{keyword}_100순위.xlsx")