from openpyxl import Workbook

# 엑셀 생성
wb = Workbook()

# 엑셀 시트 생성
ws = wb.create_sheet("startcoding")

# 셀 데이터 추가
ws["A1"] = "스타트코딩"

# 엑셀 저장
wb.save("test.xlsx")